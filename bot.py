import telebot
import subprocess
import textwrap
from PIL import Image, ImageFont, ImageDraw
from datetime import datetime

import httplib2
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from config import Config
import openpyxl.utils.cell

max_users = 99
max_rounds = 20
offset_rows = 1
offset_cols = 3

bot = telebot.TeleBot(Config.TG_TOKEN)
credentials = Credentials.from_authorized_user_file(
    "token.json",
    [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ],
)

service = build("sheets", "v4", credentials=credentials)


class UserSession:
    def __init__(self):
        self.round_id = None
        self.user_id = None
        self.score = None


sessions = {}

round_set = {}
spreadsheetId = None
create_msg = None


@bot.message_handler(content_types=["text"])
def start(message):
    global spreadsheetId
    if message.chat.id not in sessions:
        sessions[message.chat.id] = UserSession()
    if message.text == "/simple":
        simple_processor(message)
    elif message.text == "/help" or message.text == "/start":
        bot.send_message(
            message.chat.id,
            (
                "/simple - простая проверка одного квадрата\n"
                "Чтобы автоматически поддерживать таблицу нужно:\n"
                "1. Один раз запустить /create_game\n"
                "2. Выбрать на каждом проверяющем текущий раунд /round\n"
                "3. Проверить очередного участника командой /new\n"
                "4. Текущую таблицу можно узнать по команде /link"
            ),
        )
    elif message.text == "/create_game":
        sheet_processor(message)
    elif spreadsheetId == None and message.text in ["/link", "/round", "/new"]:
        bot.send_message(message.chat.id, "Для начала нужно создать игру /create_game")
    elif message.text == "/link":
        send_table(message)
    elif message.text == "/round":
        round_processor(message)
    elif message.text == "/new":
        new_processor(message)
    else:
        bot.send_message(
            message.chat.id, "Сообщение не распознано. Попробуй написать /help"
        )


def simple_processor(message):
    bot.send_message(message.chat.id, "Введи номера в квадрате")
    bot.register_next_step_handler(message, simple_game)


def simple_game(message):
    try:
        numbers = list(map(int, message.text.split()))
        if not valid_set(numbers):
            bot.send_message(message.chat.id, "Некорректный набор чисел")
        else:
            args = ["./quad"]
            for n in numbers:
                args.append(str(n))
            x = subprocess.run(args, capture_output=True, text=True)
            solution = x.stdout
            get_image_from_text(solution)
            bot.send_photo(message.chat.id, get_image_from_text(solution))
            bot.send_message(message.chat.id, "Счёт: {}".format(x.stderr))
    except Exception as exc:
        print(exc)
        bot.send_message(message.chat.id, "Что-то пошло не так, попробуй ещё раз")


def sheet_processor(message):
    global spreadsheetId, create_msg
    if spreadsheetId == None or message.text == "/yes":
        bot.send_message(message.chat.id, "Создаётся....")
        create_msg = message
        spreadsheetId = create_sheet()
        with open("/tmp/users.log", "a") as myfile:
            myfile.write(
                "Ссылка на таблицу результатов: https://docs.google.com/spreadsheets/d/{}\n".format(
                    spreadsheetId
                )
            )
        send_table(message)
    elif message.text == "/create_game":
        bot.send_message(
            message.chat.id,
            (
                "Прошлая игра была создана пользователем {}, {}. "
                "Вы действительно хотите начать новую игру? "
                "Это перезапишет результаты старой."
                "\n/yes\t/no"
            ).format(user_link(create_msg.from_user), date_convert(create_msg.date)),
            parse_mode="HTML",
        )
        send_table(message)
        bot.register_next_step_handler(message, sheet_processor)
    else:
        bot.send_message(message.chat.id, "Окей")


def round_processor(message):
    bot.send_message(
        message.chat.id, "Введи номер раунда (от 1 до {})".format(max_rounds)
    )
    bot.register_next_step_handler(message, get_round_id)


def get_round_id(message):
    global sessions
    try:
        round_id = int(message.text)
        if 1 <= round_id and round_id <= max_rounds:
            sessions[message.chat.id].round_id = round_id
            round_number_processor(message)
        else:
            bot.send_message(message.chat.id, "Некорректный ввод, попробуй ещё раз")
    except Exception as exc:
        print(exc)
        bot.send_message(message.chat.id, "Что-то пошло не так, попробуй ещё раз")


def round_number_processor(message):
    global sessions, round_set
    round_id = sessions[message.chat.id].round_id
    if round_id in round_set:
        bot.send_message(
            message.chat.id,
            "Раунд уже зарегистрирован. Хочешь изменить набор чисел?\n/yes\t /no",
        )
    else:
        bot.send_message(
            message.chat.id, "Введи 25 чисел от 1 до 13 - набор чисел в раунде"
        )
    bot.register_next_step_handler(message, get_round_numbers)


def get_round_numbers(message):
    global sessions, round_set
    round_id = sessions[message.chat.id].round_id
    try:
        if message.text == "/yes":
            bot.send_message(
                message.chat.id, "Введи 25 чисел от 1 до 13 - набор чисел в раунде"
            )
            bot.register_next_step_handler(message, get_round_numbers)
        elif message.text == "/no":
            bot.send_message(
                message.chat.id, "Отлично, раунд изменён на {}".format(round_id)
            )
        else:
            numbers = list(map(int, message.text.split()))
            if not valid_set(numbers):
                bot.send_message(
                    message.chat.id, "Некорректный набор чисел, попробуйте ещё раз"
                )
            else:
                round_set[round_id] = numbers
            bot.reply_to(
                message,
                "Отлично, набор чисел сохранён, раунд изменён на {}".format(round_id),
            )
    except Exception as exc:
        print(exc)
        bot.send_message(message.chat.id, "Что-то пошло не так, попробуй ещё раз")


def new_processor(message):
    global sessions
    if sessions[message.chat.id].round_id == None:
        bot.send_message(message.chat.id, "Раунд не выбран /round")
    else:
        bot.send_message(
            message.chat.id, "Введи номер пользователя (от 1 до {})".format(max_users)
        )
        bot.register_next_step_handler(message, get_user_id)


def get_user_id(message):
    global sessions
    try:
        user_id = int(message.text)
        if 1 <= user_id and user_id <= max_users:
            bot.send_message(
                message.chat.id,
                "Отлично. Теперь введи 25 чисел от 1 до 13 - значения квадрата",
            )
            bot.register_next_step_handler(message, get_square)
            sessions[message.chat.id].user_id = user_id
        else:
            bot.send_message(message.chat.id, "Некорректный ввод, попробуйте ещё раз")
    except Exception as exc:
        print(exc)
        bot.send_message(message.chat.id, "Что-то пошло не так, попробуй ещё раз")


def get_square(message):
    global sessions, round_set
    set = round_set[sessions[message.chat.id].round_id]
    try:
        numbers = list(map(int, message.text.split()))
        if not valid_set(numbers):
            bot.send_message(
                message.chat.id, "Некорректный набор чисел, попробуйте ещё раз"
            )
        elif sorted(numbers) != sorted(set):
            bot.send_message(
                message.chat.id,
                "Набор чисел не совпадает с набором в раунде. \nДолжен быть: {}".format(
                    "".join(map(str, sorted(set)))
                ),
            )
        else:
            args = ["./quad"]
            for n in numbers:
                args.append(str(n))
            x = subprocess.run(args, capture_output=True, text=True)
            solution = x.stdout
            get_image_from_text(solution)
            bot.send_photo(message.chat.id, get_image_from_text(solution))
            sessions[message.chat.id].score = int(x.stderr)
            bot.send_message(
                message.chat.id,
                "Счёт: {}\nПользоваель: {}\nРаунд: {}\nСохранить?\n /yes \t /no".format(
                    sessions[message.chat.id].score,
                    sessions[message.chat.id].user_id,
                    sessions[message.chat.id].round_id,
                ),
            )
            bot.register_next_step_handler(message, save_score)
    except Exception as exc:
        print(exc)
        bot.send_message(message.chat.id, "Что-то пошло не так, попробуй ещё раз")


def save_score(message):
    try:
        global sessions
        if message.text == "/yes":
            score = sessions[message.chat.id].score
            user_id = sessions[message.chat.id].user_id
            round_id = sessions[message.chat.id].round_id
            with open("/tmp/users.log", "a") as myfile:
                myfile.write(
                    "Счёт: {}\nПользоваель: {}\nРаунд: {}\n".format(
                        score, user_id, round_id
                    )
                )
            service.spreadsheets().values().update(
                spreadsheetId=spreadsheetId,
                valueInputOption="USER_ENTERED",
                range="Рабочая!{0}{1}:{0}{1}".format(
                    get_column_letter(round_id + offset_cols), user_id + offset_rows
                ),
                body={
                    "majorDimension": "ROWS",
                    "values": [[sessions[message.chat.id].score]],
                },
            ).execute()
            bot.send_message(message.chat.id, "Сохранено")
        else:
            bot.send_message(message.chat.id, "Окей, не сохраняем")
    except Exception as exc:
        print(exc)
        bot.send_message(message.chat.id, "Что-то пошло не так, попробуй ещё раз")


def get_image_from_text(text):
    img = Image.new("RGB", (400, 350), color=(225, 225, 225))
    font = ImageFont.truetype("roboto.ttf", 20)
    ImageDraw.Draw(img).multiline_text((20, 10), text, font=font, fill=(0, 0, 0))
    return img


def valid_set(numbers):
    return len(numbers) == 25 and all(1 <= x and x <= 13 for x in numbers)


def create_sheet():
    spreadsheet = (
        service.spreadsheets()
        .create(
            body={
                "properties": {"title": "Математические квадраты", "locale": "ru_RU"},
                "sheets": [
                    {
                        "properties": {
                            "sheetType": "GRID",
                            "sheetId": 0,
                            "title": "Рабочая",
                            "gridProperties": {
                                "rowCount": max_users + 2,
                                "columnCount": max_rounds + 10,
                            },
                        }
                    },
                    {
                        "properties": {
                            "sheetType": "GRID",
                            "sheetId": 1,
                            "title": "Отсортированная",
                            "gridProperties": {
                                "rowCount": max_users + 2,
                                "columnCount": max_rounds + 10,
                            },
                        }
                    },
                ],
            }
        )
        .execute()
    )
    spreadsheetId = spreadsheet["spreadsheetId"]
    print(
        "Ссылка на таблицу результатов: https://docs.google.com/spreadsheets/d/{}".format(
            spreadsheetId
        )
    )
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheetId,
        valueInputOption="USER_ENTERED",
        range="Рабочая!{0}{1}:{0}{2}".format(
            get_column_letter(0), offset_rows + 1, offset_rows + max_users
        ),
        body={
            "majorDimension": "COLUMNS",
            "values": [[i for i in range(1, max_users + 1)]],
        },
    ).execute()
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheetId,
        valueInputOption="USER_ENTERED",
        range="Рабочая!{0}{1}:{0}{2}".format(
            get_column_letter(2), offset_rows + 1, offset_rows + max_users
        ),
        body={
            "majorDimension": "COLUMNS",
            "values": [
                [
                    "=SUM({0}{1}:{1})".format(
                        get_column_letter(offset_cols), offset_rows + i
                    )
                    for i in range(1, max_users + 1)
                ]
            ],
        },
    ).execute()

    service.spreadsheets().values().update(
        spreadsheetId=spreadsheetId,
        valueInputOption="USER_ENTERED",
        range="Рабочая!{0}{1}:{2}{1}".format(
            get_column_letter(offset_cols),
            1,
            get_column_letter(offset_cols + max_rounds - 1),
        ),
        body={
            "majorDimension": "ROWS",
            "values": [[i for i in range(1, max_rounds + 1)]],
        },
    ).execute()
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheetId,
        valueInputOption="USER_ENTERED",
        range="Отсортированная!{0}{1}:{2}{1}".format(
            get_column_letter(offset_cols),
            1,
            get_column_letter(offset_cols + max_rounds - 1),
        ),
        body={
            "majorDimension": "ROWS",
            "values": [[i for i in range(1, max_rounds + 1)]],
        },
    ).execute()
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheetId,
        valueInputOption="USER_ENTERED",
        range="Рабочая!A1:C1",
        body={"majorDimension": "ROWS", "values": [["Номер", "ФИО", "Сумма"]]},
    ).execute()
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheetId,
        valueInputOption="USER_ENTERED",
        range="Отсортированная!A1:C1",
        body={"majorDimension": "ROWS", "values": [["Номер", "ФИО", "Сумма"]]},
    ).execute()
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheetId,
        valueInputOption="USER_ENTERED",
        range="Отсортированная!A{0}:A{0}".format(offset_rows + 1),
        body={
            "majorDimension": "ROWS",
            "values": [
                [
                    "=SORT(Рабочая!A{0}:{1}{2};3;FALSE)".format(
                        offset_rows + 1,
                        get_column_letter(offset_cols + max_rounds - 1),
                        offset_rows + max_users,
                    )
                ]
            ],
        },
    ).execute()
    driveservice = build("drive", "v3", credentials=credentials)
    access = (
        driveservice.permissions()
        .create(
            fileId=spreadsheetId, body={"type": "anyone", "role": "writer"}, fields="id"
        )
        .execute()
    )
    return spreadsheetId


def send_table(message):
    bot.send_message(
        message.chat.id,
        "Ссылка на таблицу результатов: https://docs.google.com/spreadsheets/d/{}".format(
            spreadsheetId
        ),
    )


def get_column_letter(num):
    return openpyxl.utils.cell.get_column_letter(num + 1)


def date_convert(ts):
    return datetime.utcfromtimestamp(ts).strftime("%d.%m.%Y %H:%M:%S")


def user_link(user):
    name = escape(user.first_name)
    return f"<a href='tg://user?id={user.id}'>{name}</a>"


def escape(text):
    chars = {"&": "&amp;", "<": "&lt;", ">": "&gt"}
    for old, new in chars.items():
        text = text.replace(old, new)
    return text


bot.infinity_polling()
